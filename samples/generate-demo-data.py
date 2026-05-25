#!/usr/bin/env python3
"""
Generate synthetic Diamant Summen-/Saldenliste sample files (SKR03).

Produces three monthly files for Q1 2026 in samples/, modeled on the
typical Diamant Excel export layout: header block with mandant +
period metadata, then a header row, then the account data, ending
with a summary row.

USE FOR DEMO / DEVELOPMENT ONLY — the values are made up but the
sums balance (EB Soll = EB Haben; VK Soll = VK Haben; Saldo Soll =
Saldo Haben) so the parser's plausibility check passes.
"""

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from pathlib import Path
import calendar

# --- SKR03 chart of accounts used in the demo ---
# (Konto-Nr, Bezeichnung, Typ)
#   AKTIVA / AUFWAND -> Saldo normalerweise im Soll
#   PASSIVA / ERTRAG -> Saldo normalerweise im Haben
KONTEN = [
    # --- Aktiva (Bilanz) ---
    ("0700", "Maschinen und maschinelle Anlagen",             "Aktiva"),
    ("0810", "Geschäftsausstattung",                          "Aktiva"),
    ("0820", "Büromaschinen, EDV-Anlagen",                    "Aktiva"),
    ("1200", "Bank",                                          "Aktiva"),
    ("1400", "Forderungen aus Lieferungen und Leistungen",    "Aktiva"),
    ("1576", "Abziehbare Vorsteuer 19 %",                     "Aktiva"),
    ("1577", "Abziehbare Vorsteuer 7 %",                      "Aktiva"),
    ("1600", "Kasse",                                         "Aktiva"),

    # --- Passiva (Bilanz) ---
    ("0640", "Darlehen, langfristig",                         "Passiva"),
    ("0800", "Gezeichnetes Kapital",                          "Passiva"),
    ("0860", "Gewinnvortrag vor Verwendung",                  "Passiva"),
    ("0900", "Gesetzliche Rücklage",                          "Passiva"),
    ("1610", "Verbindlichkeiten aus Lieferungen und Leistungen", "Passiva"),
    ("1700", "Umsatzsteuer 19 %",                             "Passiva"),
    ("1701", "Umsatzsteuer 7 %",                              "Passiva"),
    ("1740", "Verbindlichkeiten Krankenkassen",               "Passiva"),
    ("1741", "Verbindlichkeiten Lohn- und Kirchensteuer",     "Passiva"),

    # --- Erträge (GuV) ---
    ("8050", "Sonstige betriebliche Erträge",                 "Ertrag"),
    ("8100", "Zinserträge Bank",                              "Ertrag"),
    ("8200", "Erlöse steuerfrei mit Vorsteuerabzug",          "Ertrag"),
    ("8300", "Erlöse 7 % USt",                                "Ertrag"),
    ("8400", "Erlöse 19 % USt",                               "Ertrag"),

    # --- Aufwendungen (GuV) ---
    ("3300", "Wareneingang 19 % Vorsteuer",                   "Aufwand"),
    ("3400", "Wareneingang 7 % Vorsteuer",                    "Aufwand"),
    ("4100", "Löhne",                                         "Aufwand"),
    ("4110", "Gehälter",                                      "Aufwand"),
    ("4120", "Geschäftsführergehalt",                         "Aufwand"),
    ("4130", "Gesetzliche soziale Aufwendungen",              "Aufwand"),
    ("4210", "Miete für Geschäftsräume",                      "Aufwand"),
    ("4250", "Reinigung",                                     "Aufwand"),
    ("4360", "Werbeaufwand",                                  "Aufwand"),
    ("4380", "Reisekosten Arbeitnehmer",                      "Aufwand"),
    ("4500", "KFZ-Kosten",                                    "Aufwand"),
    ("4700", "Reparaturen und Instandhaltung",                "Aufwand"),
    ("4800", "Beiträge und Versicherungen",                   "Aufwand"),
    ("4830", "Abschreibungen auf Sachanlagen",                "Aufwand"),
    ("4900", "Sonstige betriebliche Aufwendungen",            "Aufwand"),
    ("4910", "Porto",                                         "Aufwand"),
    ("4920", "Telefon",                                       "Aufwand"),
    ("4930", "Strom",                                         "Aufwand"),
    ("4940", "Bürobedarf",                                    "Aufwand"),
    ("6800", "Geldverkehrskosten",                            "Aufwand"),
    ("7300", "Zinsen für kurzfristige Verbindlichkeiten",     "Aufwand"),
]

# --- Eröffnungsbilanz (zum 1.1. des Jahres, balanced Soll = Haben) ---
# Diff korrigiert via Gewinnvortrag (0860).
EB = {
    "0700": ("Soll",   220_000.00),
    "0810": ("Soll",    45_000.00),
    "0820": ("Soll",    18_500.00),
    "1200": ("Soll",   185_000.00),
    "1400": ("Soll",    78_400.00),
    "1576": ("Soll",     8_200.00),
    "1577": ("Soll",     1_100.00),
    "1600": ("Soll",     2_300.00),
    "0640": ("Haben",  148_000.00),
    "0800": ("Haben",   25_000.00),
    "0860": ("Haben",  280_500.00),   # auto-balance konto (Soll == Haben)
    "0900": ("Haben",   12_500.00),
    "1610": ("Haben",   62_450.00),
    "1700": ("Haben",   15_800.00),
    "1701": ("Haben",    1_950.00),
    "1740": ("Haben",    4_500.00),
    "1741": ("Haben",    7_800.00),
}

# --- Monthly Verkehrszahlen (will be auto-balanced via Forderungen 1400) ---
MONATE = {
    1: {
        "1200": (172_900.00, 137_400.00),
        "1400": (175_950.00, 165_200.00),
        "1576": ( 14_402.50,      0.00),
        "1577": (  2_502.50,      0.00),
        "1600": (    600.00,    410.00),
        "1700": (      0.00, 24_700.00),
        "1701": (      0.00,  2_502.50),
        "1740": (  4_350.00,  4_500.00),
        "1741": (  7_650.00,  7_800.00),
        "1610": ( 89_352.50, 91_800.00),
        "8400": (      0.00,130_000.00),
        "8300": (      0.00, 35_750.00),
        "8200": (      0.00,  4_200.00),
        "8050": (      0.00,    750.00),
        "8100": (      0.00,     50.00),
        "3300": ( 65_500.00,      0.00),
        "3400": ( 10_200.00,      0.00),
        "4100": ( 12_400.00,      0.00),
        "4110": ( 18_500.00,      0.00),
        "4120": (  8_500.00,      0.00),
        "4130": (  7_200.00,      0.00),
        "4210": (  5_000.00,      0.00),
        "4250": (    280.00,      0.00),
        "4360": (    900.00,      0.00),
        "4380": (    540.00,      0.00),
        "4500": (  1_180.00,      0.00),
        "4700": (  2_100.00,      0.00),
        "4800": (    420.00,      0.00),
        "4830": (  2_400.00,      0.00),
        "4900": (    980.00,      0.00),
        "4910": (    175.00,      0.00),
        "4920": (    310.00,      0.00),
        "4930": (    490.00,      0.00),
        "4940": (    220.00,      0.00),
        "6800": (     85.00,      0.00),
        "7300": (    210.00,      0.00),
    },
    2: {
        "1200": (158_600.00, 142_800.00),
        "1400": (162_750.00, 158_000.00),
        "1576": ( 13_847.50,      0.00),
        "1577": (  2_065.00,      0.00),
        "1600": (    520.00,    380.00),
        "1700": (      0.00, 22_800.00),
        "1701": (      0.00,  2_065.00),
        "1740": (  4_280.00,  4_350.00),
        "1741": (  7_520.00,  7_600.00),
        "1610": ( 85_100.00, 89_500.00),
        "8400": (      0.00,120_000.00),
        "8300": (      0.00, 29_500.00),
        "8200": (      0.00,  3_800.00),
        "8050": (      0.00,    620.00),
        "8100": (      0.00,     42.00),
        "3300": ( 63_000.00,      0.00),
        "3400": (  9_500.00,      0.00),
        "4100": ( 12_400.00,      0.00),
        "4110": ( 18_500.00,      0.00),
        "4120": (  8_500.00,      0.00),
        "4130": (  7_120.00,      0.00),
        "4210": (  5_000.00,      0.00),
        "4250": (    280.00,      0.00),
        "4360": (  1_150.00,      0.00),
        "4380": (    720.00,      0.00),
        "4500": (  1_220.00,      0.00),
        "4700": (    640.00,      0.00),
        "4800": (    420.00,      0.00),
        "4830": (  2_400.00,      0.00),
        "4900": (    820.00,      0.00),
        "4910": (    155.00,      0.00),
        "4920": (    310.00,      0.00),
        "4930": (    410.00,      0.00),
        "4940": (    180.00,      0.00),
        "6800": (     85.00,      0.00),
        "7300": (    195.00,      0.00),
    },
    3: {
        "1200": (181_400.00, 149_900.00),
        "1400": (180_100.00, 172_800.00),
        "1576": ( 14_877.50,      0.00),
        "1577": (  2_870.00,      0.00),
        "1600": (    700.00,    540.00),
        "1700": (      0.00, 25_175.00),
        "1701": (      0.00,  2_870.00),
        "1740": (  4_400.00,  4_500.00),
        "1741": (  7_700.00,  7_900.00),
        "1610": ( 94_440.00, 99_800.00),
        "8400": (      0.00,132_500.00),
        "8300": (      0.00, 41_000.00),
        "8200": (      0.00,  4_600.00),
        "8050": (      0.00,  1_180.00),
        "8100": (      0.00,     58.00),
        "3300": ( 68_400.00,      0.00),
        "3400": ( 10_900.00,      0.00),
        "4100": ( 12_400.00,      0.00),
        "4110": ( 19_200.00,      0.00),
        "4120": (  8_500.00,      0.00),
        "4130": (  7_400.00,      0.00),
        "4210": (  5_000.00,      0.00),
        "4250": (    280.00,      0.00),
        "4360": (  1_280.00,      0.00),
        "4380": (    810.00,      0.00),
        "4500": (  1_350.00,      0.00),
        "4700": (  3_400.00,      0.00),
        "4800": (    420.00,      0.00),
        "4830": (  2_400.00,      0.00),
        "4900": (  1_120.00,      0.00),
        "4910": (    195.00,      0.00),
        "4920": (    310.00,      0.00),
        "4930": (    480.00,      0.00),
        "4940": (    310.00,      0.00),
        "6800": (     85.00,      0.00),
        "7300": (    210.00,      0.00),
    },
}

MANDANT_NAME = "Herschel GmbH"
MANDANT_NR = "10001"
KONTENRAHMEN = "SKR03"
YEAR = 2026


def balance_monthly_data() -> None:
    """Adjust 1400 Forderungen so that each month's VK Soll = VK Haben."""
    for month, data in MONATE.items():
        soll_total = sum(s for (s, _h) in data.values())
        haben_total = sum(h for (_s, h) in data.values())
        diff = round(soll_total - haben_total, 2)
        if diff == 0:
            continue
        cur_s, cur_h = data.get("1400", (0.0, 0.0))
        if diff > 0:
            # too much Soll → add to Forderungen Haben (Einzahlung)
            data["1400"] = (cur_s, round(cur_h + diff, 2))
        else:
            # too much Haben → add to Forderungen Soll (Rechnungseingang)
            data["1400"] = (round(cur_s + (-diff), 2), cur_h)


def cumulative_through(month: int):
    """Returns cumulative (soll, haben) per account from start of year through `month`
    (inclusive), NOT counting EB. EB is added separately for balance-sheet accounts."""
    totals = {}
    for m in range(1, month + 1):
        for konto, (s, h) in MONATE.get(m, {}).items():
            cs, ch = totals.get(konto, (0.0, 0.0))
            totals[konto] = (cs + s, ch + h)
    return totals


def build_workbook(month: int) -> Workbook:
    wb = Workbook()
    ws = wb.active
    ws.title = "Summen-Salden"

    bold = Font(bold=True)
    header_fill = PatternFill(start_color="E2E8F0", end_color="E2E8F0", fill_type="solid")
    thin = Side(border_style="thin", color="CBD5E1")
    bordered = Border(top=thin, bottom=thin, left=thin, right=thin)

    last_day = calendar.monthrange(YEAR, month)[1]
    period_label = f"{month:02d}/{YEAR}"
    period_range = f"01.{month:02d}.{YEAR} - {last_day:02d}.{month:02d}.{YEAR}"

    # --- Header block (typical Diamant export style) ---
    ws["A1"] = "Mandant:"
    ws["B1"] = MANDANT_NAME
    ws["A2"] = "Mandantennummer:"
    ws["B2"] = MANDANT_NR
    ws["A3"] = "Auswertung:"
    ws["B3"] = "Summen- und Saldenliste"
    ws["A4"] = "Buchungsperiode:"
    ws["B4"] = period_label
    ws["A5"] = "Zeitraum:"
    ws["B5"] = period_range
    ws["A6"] = "Kontenrahmen:"
    ws["B6"] = KONTENRAHMEN

    for row in range(1, 7):
        ws[f"A{row}"].font = bold

    # --- Column headers in row 8 ---
    headers = [
        "Konto",
        "Bezeichnung",
        "EB Soll",
        "EB Haben",
        "Verkehrszahlen Soll",
        "Verkehrszahlen Haben",
        "Saldo Soll",
        "Saldo Haben",
    ]
    for col, h in enumerate(headers, start=1):
        c = ws.cell(row=8, column=col, value=h)
        c.font = bold
        c.fill = header_fill
        c.border = bordered
        c.alignment = Alignment(horizontal="center", vertical="center")

    # --- Data rows ---
    curr_month = MONATE.get(month, {})
    cum = cumulative_through(month)

    sum_eb_s = sum_eb_h = sum_vk_s = sum_vk_h = sum_sa_s = sum_sa_h = 0.0
    row = 9

    for konto, bezeichnung, _typ in KONTEN:
        eb_s = eb_h = 0.0
        if konto in EB:
            seite, betrag = EB[konto]
            if seite == "Soll":
                eb_s = betrag
            else:
                eb_h = betrag

        vk_s, vk_h = curr_month.get(konto, (0.0, 0.0))
        cum_s, cum_h = cum.get(konto, (0.0, 0.0))

        # Saldo = wohin der Netto-Effekt geht.
        total_soll = eb_s + cum_s
        total_haben = eb_h + cum_h
        if total_soll > total_haben:
            saldo_s = round(total_soll - total_haben, 2)
            saldo_h = 0.0
        else:
            saldo_s = 0.0
            saldo_h = round(total_haben - total_soll, 2)

        ws.cell(row=row, column=1, value=konto)
        ws.cell(row=row, column=2, value=bezeichnung)
        ws.cell(row=row, column=3, value=round(eb_s, 2) if eb_s else None)
        ws.cell(row=row, column=4, value=round(eb_h, 2) if eb_h else None)
        ws.cell(row=row, column=5, value=round(vk_s, 2) if vk_s else None)
        ws.cell(row=row, column=6, value=round(vk_h, 2) if vk_h else None)
        ws.cell(row=row, column=7, value=saldo_s if saldo_s else None)
        ws.cell(row=row, column=8, value=saldo_h if saldo_h else None)

        for col in range(3, 9):
            cell = ws.cell(row=row, column=col)
            cell.number_format = "#,##0.00"
            cell.alignment = Alignment(horizontal="right")

        sum_eb_s += eb_s
        sum_eb_h += eb_h
        sum_vk_s += vk_s
        sum_vk_h += vk_h
        sum_sa_s += saldo_s
        sum_sa_h += saldo_h

        row += 1

    # --- Summen-Zeile ---
    ws.cell(row=row, column=2, value="SUMMEN").font = bold
    ws.cell(row=row, column=3, value=round(sum_eb_s, 2)).font = bold
    ws.cell(row=row, column=4, value=round(sum_eb_h, 2)).font = bold
    ws.cell(row=row, column=5, value=round(sum_vk_s, 2)).font = bold
    ws.cell(row=row, column=6, value=round(sum_vk_h, 2)).font = bold
    ws.cell(row=row, column=7, value=round(sum_sa_s, 2)).font = bold
    ws.cell(row=row, column=8, value=round(sum_sa_h, 2)).font = bold
    for col in range(3, 9):
        cell = ws.cell(row=row, column=col)
        cell.number_format = "#,##0.00"
        cell.alignment = Alignment(horizontal="right")
        cell.border = bordered

    # --- Column widths ---
    widths = [10, 48, 14, 14, 18, 20, 14, 14]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    return wb


def main():
    out_dir = Path(__file__).parent
    balance_monthly_data()
    for month in (1, 2, 3):
        wb = build_workbook(month)
        fname = out_dir / f"diamant-summenliste-{YEAR}-{month:02d}.xlsx"
        wb.save(fname)
        print(f"  wrote {fname.relative_to(out_dir.parent)}")


if __name__ == "__main__":
    main()
